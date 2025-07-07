import streamlit as st
import pandas as pd
from io import BytesIO
import re
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, PageBreak, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from collections import defaultdict
from reportlab.lib.units import mm
from reportlab.lib.pagesizes import letter
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import io
import streamlit.components.v1 as components
from openpyxl.utils import get_column_letter
import time
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import pytz

st.set_page_config(
    page_title="OSG DASHBOARD",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Force light mode and disable dark mode
st.markdown("""
<style>
/* Define light and dark themes explicitly */
html {
  --primary-light: #3498db;
  --secondary-light: #2980b9;
  --text-light: #2c3e50;
  --bg-light: #ffffff;
  --card-bg-light: #f8f9fa;
  --border-light: #dfe6e9;

  --primary-dark: #2980b9;
  --secondary-dark: #1c5d99;
  --text-dark: #ecf0f1;
  --bg-dark: #1e293b;
  --card-bg-dark: #334155;
  --border-dark: #475569;
}

/* Set default (light) mode */
body {
  background-color: var(--bg-light);
  color: var(--text-light);
}

/* Dark mode override */
@media (prefers-color-scheme: dark) {
  body {
    background-color: var(--bg-dark);
    color: var(--text-dark);
  }
}

/* Card Styling */
.report-card {
    border-radius: 12px;
    padding: 1.5rem;
    margin-bottom: 1.5rem;
    box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1);
    border-left: 4px solid var(--primary-light);
    background-color: var(--card-bg-light);
}

@media (prefers-color-scheme: dark) {
    .report-card {
        background-color: var(--card-bg-dark);
        border-left: 4px solid var(--primary-dark);
        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.4);
    }
}

/* Title Styling */
.report-title {
    font-size: 1.75rem;
    font-weight: 700;
    margin-bottom: 0.5rem;
    border-bottom: 2px solid var(--primary-light);
    padding-bottom: 0.5rem;
    color: var(--text-light);
}

@media (prefers-color-scheme: dark) {
    .report-title {
        color: var(--text-dark);
        border-bottom: 2px solid var(--primary-dark);
    }
}

/* Subtitle */
.report-subtitle {
    font-size: 1.25rem;
    font-weight: 600;
    margin: 1rem 0 0.5rem 0;
    color: var(--text-light);
}

@media (prefers-color-scheme: dark) {
    .report-subtitle {
        color: var(--text-dark);
    }
}

/* Time Indicator */
.time-indicator {
    display: inline-block;
    background-color: var(--primary-light);
    color: white;
    padding: 0.25rem 0.75rem;
    border-radius: 20px;
    font-size: 0.9rem;
    font-weight: 500;
}

@media (prefers-color-scheme: dark) {
    .time-indicator {
        background-color: var(--primary-dark);
    }
}

/* File Uploader */
.stFileUploader > div > div {
    border: 2px dashed var(--border-light);
    border-radius: 12px;
    padding: 2rem;
    background-color: var(--card-bg-light);
    transition: all 0.3s ease;
}

.stFileUploader > div > div:hover {
    border-color: var(--primary-light);
    background-color: rgba(52, 152, 219, 0.05);
}

@media (prefers-color-scheme: dark) {
    .stFileUploader > div > div {
        border: 2px dashed var(--border-dark);
        background-color: var(--card-bg-dark);
    }
    .stFileUploader > div > div:hover {
        border-color: var(--primary-dark);
        background-color: rgba(41, 128, 185, 0.1);
    }
}

/* Default File Message */
.default-file {
    font-size: 0.9rem;
    margin-top: 1rem;
    padding: 0.75rem;
    border-radius: 8px;
    border-left: 3px solid var(--primary-light);
    background-color: rgba(52, 152, 219, 0.1);
    color: var(--text-light);
}

@media (prefers-color-scheme: dark) {
    .default-file {
        background-color: rgba(41, 128, 185, 0.2);
        border-left: 3px solid var(--primary-dark);
        color: var(--text-dark);
    }
}

.info-box {
    background-color: #e8f4f8;
    border: 1px solid #bee5eb;
    border-radius: 8px;
    padding: 1rem;
    margin: 1rem 0;
}

.header {
    color: #2c3e50;
    font-size: 2rem;
    font-weight: bold;
    margin-bottom: 1rem;
}

.insight-box {
    background-color: #f8f9fa;
    padding: 15px;
    border-radius: 10px;
    border-left: 4px solid #007bff;
    margin: 10px 0;
}

.insight-box h4 {
    color: #007bff;
    margin-top: 0;
}

.insight-box ul {
    margin-bottom: 0;
}

.insight-box li {
    margin: 5px 0;
}
</style>
""", unsafe_allow_html=True)

# Neon glowing icons as SVG for tabs
tab_icons = {
    "📊 OSG REPORT 1": """<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" fill="currentColor"><path d="M13 2h-2v10h2V2zM6 9h2v13H6V9zm10 0h2v13h-2V9z"/></svg>""",
    "📊 OSG REPORT 2": """<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" fill="currentColor"><path d="M3 17h2v4H3v-4zm4-6h2v10H7V11zm4-4h2v14h-2V7zm4 6h2v8h-2v-8z"/></svg>""",
    "🔗 Data Mapping": """<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" fill="currentColor"><path d="M15.5 14h-.79l-.28-.27A6.471 6.471 0 0016 9.5 6.5 6.5 0 109.5 16c1.61 0 3.09-.59 4.23-1.57l.27.28v.79l5 4.99L20.49 19l-4.99-5zM9.5 14C7.57 14 6 12.43 6 10.5S7.57 7 9.5 7 13 8.57 13 10.5 11.43 14 9.5 14z"/></svg>"""
}

# Streamlit Tabs with icons + neon styles
tab1, tab2, tab3 = st.tabs(list(tab_icons.keys()))

# --------------------------- REPORT 1 TAB ---------------------------
with tab1:
    st.markdown('<h1 class="header">OSG All Store Report with Conversion Rates</h1>', unsafe_allow_html=True)

    st.markdown("""
        <div class="info-box">
            <strong>Instructions:</strong> Upload the following files to generate the comprehensive sales report:
            <ul>
                <li><strong>Current Month OSG Sales Data:</strong> Contains OSG sales data (Store, DATE, QUANTITY, AMOUNT)</li>
                <li><strong>Previous Month OSG Sales Data:</strong> Contains previous month OSG sales data (Store, DATE, QUANTITY, AMOUNT)</li>
                <li><strong>Current Month Product Sales Data:</strong> Contains product sales data (Date, Branch, Sold Price)</li>
                <li><strong>myG All Store List:</strong> Loaded by default</li>
                <li><strong>Store, RBM List:</strong> Loaded by default</li>
            </ul>
            <strong>Note:</strong> Conversion Rate = (OSG Sale / Product Sale) × 100
        </div>
        """, unsafe_allow_html=True)

    col1, col2 = st.columns(2)
    with col1:
        report_date = st.date_input("Select current report date", value=datetime.today())
    with col2:
        prev_date = st.date_input("Select previous report date (for comparison)", value=datetime.today().replace(day=1))

    # File uploads
    book1_file = st.file_uploader("Upload Current Month OSG Sales Data", type=["xlsx"], key="curr_sales")
    prev_month_file = st.file_uploader("Upload Previous Month OSG Sales Data", type=["xlsx"], key="prev_sales")
    product_file = st.file_uploader("Upload Current Month Product Sales Data", type=["xlsx"], key="product_sales")

    store_list_file = "myG All Store.xlsx"
    rbm_file = "RBM,BDM,BRANCH.xlsx"

    try:
        future_store_df = pd.read_excel(store_list_file)
        rbm_df = pd.read_excel(rbm_file)
        st.success("✅ Loaded default myG All Store List & Store, RBM List.")
    except Exception as e:
        st.error(f"Error loading defaults: {e}")
        st.stop()

    if book1_file and product_file:
        with st.spinner('Processing data...'):
            # Load and preprocess OSG sales data
            book1_df = pd.read_excel(book1_file)
            book1_df.rename(columns={'Branch': 'Store'}, inplace=True)
            book1_df['DATE'] = pd.to_datetime(book1_df['DATE'], dayfirst=True, errors='coerce')
            book1_df = book1_df.dropna(subset=['DATE'])
            rbm_df.rename(columns={'Branch': 'Store'}, inplace=True)

            # Load and preprocess product sales data
            product_df = pd.read_excel(product_file)
            product_df.rename(columns={'Branch': 'Store', 'Date': 'DATE', 'Sold Price': 'AMOUNT'}, inplace=True)
            product_df['DATE'] = pd.to_datetime(product_df['DATE'], dayfirst=True, errors='coerce')
            product_df = product_df.dropna(subset=['DATE'])
            if 'QUANTITY' not in product_df.columns:
                product_df['QUANTITY'] = 1  # Assume 1 unit per sale if QUANTITY is not provided

            today = pd.to_datetime(report_date)
            # Process OSG data
            mtd_df = book1_df[book1_df['DATE'].dt.month == today.month]
            today_df = mtd_df[mtd_df['DATE'].dt.date == today.date()]
            today_agg = today_df.groupby('Store', as_index=False).agg({'QUANTITY': 'sum', 'AMOUNT': 'sum'}).rename(columns={'QUANTITY': 'FTD Count', 'AMOUNT': 'FTD Value'})
            mtd_agg = mtd_df.groupby('Store', as_index=False).agg({'QUANTITY': 'sum', 'AMOUNT': 'sum'}).rename(columns={'QUANTITY': 'MTD Count', 'AMOUNT': 'MTD Value'})

            # Process product data
            product_mtd_df = product_df[product_df['DATE'].dt.month == today.month]
            product_today_df = product_mtd_df[product_mtd_df['DATE'].dt.date == today.date()]
            product_today_agg = product_today_df.groupby('Store', as_index=False).agg({'QUANTITY': 'sum', 'AMOUNT': 'sum'}).rename(columns={'QUANTITY': 'Product_FTD_Count', 'AMOUNT': 'Product_FTD_Amount'})
            product_mtd_agg = product_mtd_df.groupby('Store', as_index=False).agg({'QUANTITY': 'sum', 'AMOUNT': 'sum'}).rename(columns={'QUANTITY': 'Product_MTD_Count', 'AMOUNT': 'Product_MTD_Amount'})

            # Process previous month data if provided
            if prev_month_file:
                prev_df = pd.read_excel(prev_month_file)
                prev_df.rename(columns={'Branch': 'Store'}, inplace=True)
                prev_df['DATE'] = pd.to_datetime(prev_df['DATE'], dayfirst=True, errors='coerce')
                prev_df = prev_df.dropna(subset=['DATE'])
                prev_month = pd.to_datetime(prev_date)
                prev_mtd_df = prev_df[prev_df['DATE'].dt.month == prev_month.month]
                prev_mtd_agg = prev_mtd_df.groupby('Store', as_index=False).agg({'AMOUNT': 'sum'}).rename(columns={'AMOUNT': 'PREV MONTH SALE'})
            else:
                prev_mtd_agg = pd.DataFrame(columns=['Store', 'PREV MONTH SALE'])

            # Merge all data
            all_stores = pd.DataFrame(pd.Series(pd.concat([future_store_df['Store'], book1_df['Store'], product_df['Store']]).unique(), name='Store'))
            report_df = all_stores.merge(today_agg, on='Store', how='left') \
                                  .merge(mtd_agg, on='Store', how='left') \
                                  .merge(product_today_agg, on='Store', how='left') \
                                  .merge(product_mtd_agg, on='Store', how='left') \
                                  .merge(prev_mtd_agg, on='Store', how='left') \
                                  .merge(rbm_df[['Store', 'RBM']], on='Store', how='left')

            # Ensure all required columns exist
            required_columns = ['Store', 'FTD Count', 'FTD Value', 'Product_FTD_Amount', 'MTD Count', 'MTD Value', 'Product_MTD_Amount', 'PREV MONTH SALE', 'RBM']
            for col in required_columns:
                if col not in report_df.columns:
                    report_df[col] = 0
            report_df = report_df.rename(columns={'Store': 'Store Name'})  # Rename 'Store' to 'Store Name'

            # Fill NaN values and ensure integer types
            report_df[['FTD Count', 'FTD Value', 'MTD Count', 'MTD Value', 'Product_FTD_Count', 'Product_FTD_Amount', 'Product_MTD_Count', 'Product_MTD_Amount', 'PREV MONTH SALE']] = report_df[['FTD Count', 'FTD Value', 'MTD Count', 'MTD Value', 'Product_FTD_Count', 'Product_FTD_Amount', 'Product_MTD_Count', 'Product_MTD_Amount', 'PREV MONTH SALE']].fillna(0).astype(int)

            # Calculate difference percentage
            report_df['DIFF %'] = report_df.apply(
                lambda x: round(((x['MTD Value'] - x['PREV MONTH SALE']) / x['PREV MONTH SALE']) * 100, 2) if x['PREV MONTH SALE'] != 0 else 0,
                axis=1
            )

            # Calculate ASP
            report_df['ASP'] = report_df.apply(
                lambda x: round(x['MTD Value'] / x['MTD Count'], 2) if x['MTD Count'] != 0 else 0,
                axis=1
            )

            # Calculate conversion rates
            report_df['FTD Value Conversion'] = report_df.apply(
                lambda x: round((x['FTD Value'] / x['Product_FTD_Amount']) * 100, 2) if x['Product_FTD_Amount'] != 0 else 0,
                axis=1
            )
            report_df['MTD Value Conversion'] = report_df.apply(
                lambda x: round((x['MTD Value'] / x['Product_MTD_Amount']) * 100, 2) if x['Product_MTD_Amount'] != 0 else 0,
                axis=1
            )

            # Excel report generation
            excel_output = BytesIO()
            with pd.ExcelWriter(excel_output, engine='xlsxwriter') as writer:
                workbook = writer.book

                colors_palette = {
                    'primary_blue': '#1E3A8A',
                    'light_blue': '#DBEAFE',
                    'success_green': '#065F46',
                    'light_green': '#D1FAE5',
                    'warning_orange': '#EA580C',
                    'light_orange': '#FED7AA',
                    'danger_red': '#DC2626',
                    'light_red': '#FEE2E2',
                    'accent_purple': '#7C3AED',
                    'light_purple': '#EDE9FE',
                    'neutral_gray': '#6B7280',
                    'light_gray': '#F9FAFB',
                    'white': '#FFFFFF',
                    'dark_blue': '#0F172A',
                    'mint_green': '#10B981',
                    'light_mint': '#ECFDF5',
                    'royal_blue': '#3B82F6',
                    'light_royal': '#EBF8FF'
                }

                formats = {
                    'title': workbook.add_format({
                        'bold': True, 'font_size': 16, 'font_color': colors_palette['primary_blue'],
                        'align': 'center', 'valign': 'vcenter', 'bg_color': colors_palette['white'],
                        'border': 1, 'border_color': colors_palette['primary_blue']
                    }),
                    'subtitle': workbook.add_format({
                        'bold': True, 'font_size': 12, 'font_color': colors_palette['neutral_gray'],
                        'align': 'center', 'valign': 'vcenter', 'italic': True
                    }),
                    'header_main': workbook.add_format({
                        'bold': True, 'font_size': 11, 'font_color': colors_palette['white'],
                        'bg_color': colors_palette['primary_blue'], 'align': 'center', 'valign': 'vcenter',
                        'border': 1, 'border_color': colors_palette['primary_blue'], 'text_wrap': True
                    }),
                    'header_secondary': workbook.add_format({
                        'bold': True, 'font_size': 10, 'font_color': colors_palette['primary_blue'],
                        'bg_color': colors_palette['light_blue'], 'align': 'center', 'valign': 'vcenter',
                        'border': 1, 'border_color': colors_palette['primary_blue']
                    }),
                    'data_normal': workbook.add_format({
                        'font_size': 10, 'align': 'center', 'valign': 'vcenter',
                        'border': 1, 'border_color': colors_palette['neutral_gray'], 'bg_color': colors_palette['white']
                    }),
                    'data_alternate': workbook.add_format({
                        'font_size': 10, 'align': 'center', 'valign': 'vcenter',
                        'border': 1, 'border_color': colors_palette['neutral_gray'], 'bg_color': colors_palette['light_gray']
                    }),
                    'data_store_name': workbook.add_format({
                        'font_size': 10, 'bold': True, 'align': 'left', 'valign': 'vcenter',
                        'border': 1, 'border_color': colors_palette['neutral_gray'], 'bg_color': colors_palette['white'], 'indent': 1
                    }),
                    'data_store_name_alt': workbook.add_format({
                        'font_size': 10, 'bold': True, 'align': 'left', 'valign': 'vcenter',
                        'border': 1, 'border_color': colors_palette['neutral_gray'], 'bg_color': colors_palette['light_gray'], 'indent': 1
                    }),
                    'conversion_low': workbook.add_format({
                        'font_size': 10, 'font_color': colors_palette['danger_red'], 'bg_color': colors_palette['light_red'],
                        'align': 'center', 'valign': 'vcenter', 'border': 1, 'border_color': colors_palette['danger_red'], 'num_format': '0.00%', 'bold': True
                    }),
                    'conversion_format': workbook.add_format({
                        'font_size': 10, 'align': 'center', 'valign': 'vcenter',
                        'border': 1, 'border_color': colors_palette['neutral_gray'], 'num_format': '0.00%'
                    }),
                    'conversion_format_alt': workbook.add_format({
                        'font_size': 10, 'align': 'center', 'valign': 'vcenter',
                        'border': 1, 'border_color': colors_palette['neutral_gray'], 'bg_color': colors_palette['light_gray'], 'num_format': '0.00%'
                    }),
                    'total_row': workbook.add_format({
                        'bold': True, 'font_size': 11, 'font_color': colors_palette['white'],
                        'bg_color': colors_palette['accent_purple'], 'align': 'center', 'valign': 'vcenter',
                        'border': 2, 'border_color': colors_palette['accent_purple'], 'num_format': '0.00'
                    }),
                    'total_label': workbook.add_format({
                        'bold': True, 'font_size': 11, 'font_color': colors_palette['white'],
                        'bg_color': colors_palette['accent_purple'], 'align': 'center', 'valign': 'vcenter',
                        'border': 2, 'border_color': colors_palette['accent_purple']
                    }),
                    'rbm_title': workbook.add_format({
                        'bold': True, 'font_size': 18, 'font_color': colors_palette['white'],
                        'bg_color': colors_palette['dark_blue'], 'align': 'center', 'valign': 'vcenter',
                        'border': 2, 'border_color': colors_palette['dark_blue']
                    }),
                    'rbm_subtitle': workbook.add_format({
                        'bold': True, 'font_size': 11, 'font_color': colors_palette['dark_blue'],
                        'bg_color': colors_palette['light_royal'], 'align': 'center', 'valign': 'vcenter',
                        'border': 1, 'border_color': colors_palette['royal_blue'], 'italic': True
                    }),
                    'rbm_header': workbook.add_format({
                        'bold': True, 'font_size': 11, 'font_color': colors_palette['white'],
                        'bg_color': colors_palette['royal_blue'], 'align': 'center', 'valign': 'vcenter',
                        'border': 1, 'border_color': colors_palette['royal_blue'], 'text_wrap': True
                    }),
                    'rbm_data_normal': workbook.add_format({
                        'font_size': 10, 'align': 'center', 'valign': 'vcenter',
                        'border': 1, 'border_color': colors_palette['neutral_gray'], 'bg_color': colors_palette['white']
                    }),
                    'rbm_data_alternate': workbook.add_format({
                        'font_size': 10, 'align': 'center', 'valign': 'vcenter',
                        'border': 1, 'border_color': colors_palette['neutral_gray'], 'bg_color': colors_palette['light_royal']
                    }),
                    'rbm_store_name': workbook.add_format({
                        'font_size': 10, 'bold': True, 'align': 'left', 'valign': 'vcenter',
                        'border': 1, 'border_color': colors_palette['neutral_gray'], 'bg_color': colors_palette['white'], 'indent': 1
                    }),
                    'rbm_store_name_alt': workbook.add_format({
                        'font_size': 10, 'bold': True, 'align': 'left', 'valign': 'vcenter',
                        'border': 1, 'border_color': colors_palette['neutral_gray'], 'bg_color': colors_palette['light_royal'], 'indent': 1
                    }),
                    'rbm_conversion_low': workbook.add_format({
                        'font_size': 10, 'font_color': colors_palette['danger_red'], 'bg_color': colors_palette['light_red'],
                        'align': 'center', 'valign': 'vcenter', 'border': 1, 'border_color': colors_palette['danger_red'], 'num_format': '0.00%', 'bold': True
                    }),
                    'rbm_conversion_format': workbook.add_format({
                        'font_size': 10, 'align': 'center', 'valign': 'vcenter',
                        'border': 1, 'border_color': colors_palette['neutral_gray'], 'num_format': '0.00%'
                    }),
                    'rbm_conversion_format_alt': workbook.add_format({
                        'font_size': 10, 'align': 'center', 'valign': 'vcenter',
                        'border': 1, 'border_color': colors_palette['neutral_gray'], 'bg_color': colors_palette['light_royal'], 'num_format': '0.00%'
                    }),
                    'rbm_total': workbook.add_format({
                        'bold': True, 'font_size': 12, 'font_color': colors_palette['white'],
                        'bg_color': colors_palette['mint_green'], 'align': 'center', 'valign': 'vcenter',
                        'border': 2, 'border_color': colors_palette['mint_green'], 'num_format': '0.00%'
                    }),
                    'rbm_total_label': workbook.add_format({
                        'bold': True, 'font_size': 12, 'font_color': colors_palette['white'],
                        'bg_color': colors_palette['mint_green'], 'align': 'center', 'valign': 'vcenter',
                        'border': 2, 'border_color': colors_palette['mint_green']
                    }),
                    'rbm_summary': workbook.add_format({
                        'bold': True, 'font_size': 10, 'font_color': colors_palette['royal_blue'],
                        'bg_color': colors_palette['light_royal'], 'align': 'center', 'valign': 'vcenter',
                        'border': 1, 'border_color': colors_palette['royal_blue']
                    }),
                    'rbm_performance': workbook.add_format({
                        'bold': True, 'font_size': 10, 'font_color': colors_palette['white'],
                        'bg_color': colors_palette['accent_purple'], 'align': 'center', 'valign': 'vcenter',
                        'border': 1, 'border_color': colors_palette['accent_purple']
                    }),
                    'asp_format': workbook.add_format({
                        'font_size': 10, 'align': 'center', 'valign': 'vcenter',
                        'border': 1, 'border_color': colors_palette['neutral_gray'], 'num_format': '₹#,##0.00'
                    }),
                    'asp_format_alt': workbook.add_format({
                        'font_size': 10, 'align': 'center', 'valign': 'vcenter',
                        'border': 1, 'border_color': colors_palette['neutral_gray'], 'bg_color': colors_palette['light_royal'], 'num_format': '₹#,##0.00'
                    }),
                    'asp_total': workbook.add_format({
                        'bold': True, 'font_size': 12, 'font_color': colors_palette['white'],
                        'bg_color': colors_palette['mint_green'], 'align': 'center', 'valign': 'vcenter',
                        'border': 2, 'border_color': colors_palette['mint_green'], 'num_format': '₹#,##0.00'
                    })
                }

                # Set IST timezone
                ist = pytz.timezone('Asia/Kolkata')
                ist_time = datetime.now(ist)

                # ALL STORES SHEET
                all_data = report_df.sort_values('MTD Value', ascending=False)
                worksheet = workbook.add_worksheet("All Stores")

                # Dynamically adjust column widths with error handling
                headers = ['Store Name', 'FTD Count', 'FTD Value', 'FTD Value Conversion', 'MTD Count', 'MTD Value', 'MTD Value Conversion', 'PREV MONTH SALE', 'DIFF %', 'ASP']
                column_widths = {}
                for i in range(len(headers)):
                    try:
                        if i == 0:
                            max_len = max(all_data[headers[i]].astype(str).map(len).max(), len(headers[i])) + 2
                        else:
                            max_len = max(all_data[headers[i]].map(lambda x: len(str(x))).max() if headers[i] in all_data.columns else 0, len(headers[i])) + 2
                        column_widths[i] = max(max_len, 10)
                    except KeyError:
                        column_widths[i] = len(headers[i]) + 2  # Default width if column is missing
                    worksheet.set_column(i, i, column_widths[i])

                worksheet.merge_range(0, 0, 0, len(headers) - 1, "OSG All Stores Sales Report", formats['title'])
                worksheet.merge_range(1, 0, 1, len(headers) - 1, f"Report Generated: {ist_time.strftime('%d %B %Y %I:%M %p IST')}", formats['subtitle'])

                total_stores = len(all_data)
                active_stores = len(all_data[all_data['FTD Count'] > 0])
                inactive_stores = total_stores - active_stores

                worksheet.merge_range(3, 0, 3, 1, "📊 SUMMARY", formats['header_secondary'])
                worksheet.merge_range(3, 2, 3, len(headers) - 1, f"Total: {total_stores} | Active: {active_stores} | Inactive: {inactive_stores}", formats['data_normal'])

                # Headers
                for col, header in enumerate(headers):
                    worksheet.write(5, col, header, formats['header_main'])

                for row_idx, (_, row) in enumerate(all_data.iterrows(), start=6):
                    is_alternate = (row_idx - 6) % 2 == 1
                    store_format = formats['data_store_name_alt'] if is_alternate else formats['data_store_name']
                    worksheet.write(row_idx, 0, row['Store Name'], store_format)

                    data_format = formats['data_alternate'] if is_alternate else formats['data_normal']
                    worksheet.write(row_idx, 1, int(row['FTD Count']), data_format)
                    worksheet.write(row_idx, 2, int(row['FTD Value']), data_format)

                    ftd_conversion = row['FTD Value Conversion']
                    conversion_format = formats['conversion_format_alt'] if is_alternate else formats['conversion_format']
                    if ftd_conversion < 2:
                        worksheet.write(row_idx, 3, ftd_conversion / 100, formats['conversion_low'])
                    else:
                        worksheet.write(row_idx, 3, ftd_conversion / 100, conversion_format)

                    worksheet.write(row_idx, 4, int(row['MTD Count']), data_format)
                    worksheet.write(row_idx, 5, int(row['MTD Value']), data_format)

                    mtd_conversion = row['MTD Value Conversion']
                    if mtd_conversion < 2:
                        worksheet.write(row_idx, 6, mtd_conversion / 100, formats['conversion_low'])
                    else:
                        worksheet.write(row_idx, 6, mtd_conversion / 100, conversion_format)

                    worksheet.write(row_idx, 7, int(row['PREV MONTH SALE']), data_format)
                    worksheet.write(row_idx, 8, f"{row['DIFF %']}%", data_format)

                    asp_format = formats['asp_format_alt'] if is_alternate else formats['asp_format']
                    worksheet.write(row_idx, 9, row['ASP'], asp_format)

                total_row = len(all_data) + 7
                worksheet.write(total_row, 0, '🎯 TOTAL', formats['total_label'])
                worksheet.write(total_row, 1, all_data['FTD Count'].sum(), formats['total_row'])
                worksheet.write(total_row, 2, all_data['FTD Value'].sum(), formats['total_row'])
                total_ftd_conversion = round((all_data['FTD Value'].sum() / all_data['Product_FTD_Amount'].sum()) * 100, 2) if all_data['Product_FTD_Amount'].sum() != 0 else 0
                worksheet.write(total_row, 3, total_ftd_conversion / 100, formats['conversion_low'] if total_ftd_conversion < 2 else formats['total_row'])
                worksheet.write(total_row, 4, all_data['MTD Count'].sum(), formats['total_row'])
                worksheet.write(total_row, 5, all_data['MTD Value'].sum(), formats['total_row'])
                total_mtd_conversion = round((all_data['MTD Value'].sum() / all_data['Product_MTD_Amount'].sum()) * 100, 2) if all_data['Product_MTD_Amount'].sum() != 0 else 0
                worksheet.write(total_row, 6, total_mtd_conversion / 100, formats['conversion_low'] if total_mtd_conversion < 2 else formats['total_row'])
                worksheet.write(total_row, 7, all_data['PREV MONTH SALE'].sum(), formats['total_row'])
                total_diff = round(((all_data['MTD Value'].sum() - all_data['PREV MONTH SALE'].sum()) / all_data['PREV MONTH SALE'].sum()) * 100, 2) if all_data['PREV MONTH SALE'].sum() != 0 else 0
                worksheet.write(total_row, 8, f"{total_diff}%", formats['total_row'])
                total_asp = round(all_data['MTD Value'].sum() / all_data['MTD Count'].sum(), 2) if all_data['MTD Count'].sum() != 0 else 0
                worksheet.write(total_row, 9, total_asp, formats['asp_total'])

                if len(all_data) > 0:
                    top_performer = all_data.iloc[0]
                    insights_row = total_row + 2
                    worksheet.merge_range(insights_row, 0, insights_row, len(headers) - 1,
                                        f"🏆 Top Performer: {top_performer['Store Name']} (₹{int(top_performer['MTD Value']):,})",
                                        formats['data_normal'])

                # RBM SHEETS
                for rbm in report_df['RBM'].dropna().unique():
                    rbm_data = report_df[report_df['RBM'] == rbm].sort_values('MTD Value', ascending=False)
                    worksheet_name = rbm[:31] if len(rbm) > 31 else rbm
                    rbm_ws = workbook.add_worksheet(worksheet_name)

                    # Dynamically adjust column widths for RBM sheets
                    rbm_column_widths = {}
                    for i in range(len(headers)):
                        try:
                            if i == 0:
                                max_len = max(rbm_data[headers[i]].astype(str).map(len).max(), len(headers[i])) + 2
                            else:
                                max_len = max(rbm_data[headers[i]].map(lambda x: len(str(x))).max() if headers[i] in rbm_data.columns else 0, len(headers[i])) + 2
                            rbm_column_widths[i] = max(max_len, 10)
                        except KeyError:
                            rbm_column_widths[i] = len(headers[i]) + 2
                        rbm_ws.set_column(i, i, rbm_column_widths[i])

                    rbm_ws.merge_range(0, 0, 0, len(headers) - 1, f" {rbm} - Sales Performance Report", formats['rbm_title'])
                    rbm_ws.merge_range(1, 0, 1, len(headers) - 1, f"Report Period: {ist_time.strftime('%B %Y')} | Generated: {ist_time.strftime('%d %B %Y %I:%M %p IST')}", formats['rbm_subtitle'])

                    rbm_total_stores = len(rbm_data)
                    rbm_active_stores = len(rbm_data[rbm_data['FTD Count'] > 0])
                    rbm_inactive_stores = rbm_total_stores - rbm_active_stores
                    rbm_total_amount = rbm_data['MTD Value'].sum()

                    rbm_ws.merge_range(3, 0, 3, 1, "📈 PERFORMANCE OVERVIEW", formats['rbm_summary'])
                    rbm_ws.merge_range(3, 2, 3, len(headers) - 1, f"Total Stores: {rbm_total_stores} | Active: {rbm_active_stores} | Inactive: {rbm_inactive_stores} | Total Revenue: ₹{rbm_total_amount:,}", formats['rbm_summary'])

                    if len(rbm_data) > 0:
                        best_performer = rbm_data.iloc[0]
                        rbm_ws.merge_range(4, 0, 4, len(headers) - 1, f"🥇 Best Performer: {best_performer['Store Name']} - ₹{int(best_performer['MTD Value']):,}", formats['rbm_performance'])

                    # Headers
                    for col, header in enumerate(headers):
                        rbm_ws.write(6, col, header, formats['rbm_header'])

                    for row_idx, (_, row) in enumerate(rbm_data.iterrows(), start=7):
                        is_alternate = (row_idx - 7) % 2 == 1
                        store_format = formats['rbm_store_name_alt'] if is_alternate else formats['rbm_store_name']
                        rbm_ws.write(row_idx, 0, row['Store Name'], store_format)

                        data_format = formats['rbm_data_alternate'] if is_alternate else formats['rbm_data_normal']
                        rbm_ws.write(row_idx, 1, int(row['FTD Count']), data_format)
                        rbm_ws.write(row_idx, 2, int(row['FTD Value']), data_format)

                        ftd_conversion = row['FTD Value Conversion']
                        conversion_format = formats['rbm_conversion_format_alt'] if is_alternate else formats['rbm_conversion_format']
                        if ftd_conversion < 2:
                            rbm_ws.write(row_idx, 3, ftd_conversion / 100, formats['rbm_conversion_low'])
                        else:
                            rbm_ws.write(row_idx, 3, ftd_conversion / 100, conversion_format)

                        rbm_ws.write(row_idx, 4, int(row['MTD Count']), data_format)
                        rbm_ws.write(row_idx, 5, int(row['MTD Value']), data_format)

                        mtd_conversion = row['MTD Value Conversion']
                        if mtd_conversion < 2:
                            rbm_ws.write(row_idx, 6, mtd_conversion / 100, formats['rbm_conversion_low'])
                        else:
                            rbm_ws.write(row_idx, 6, mtd_conversion / 100, conversion_format)

                        rbm_ws.write(row_idx, 7, int(row['PREV MONTH SALE']), data_format)
                        rbm_ws.write(row_idx, 8, f"{row['DIFF %']}%", data_format)

                        asp_format = formats['asp_format_alt'] if is_alternate else formats['asp_format']
                        rbm_ws.write(row_idx, 9, row['ASP'], asp_format)

                    total_row = len(rbm_data) + 8
                    rbm_ws.write(total_row, 0, '🎯 TOTAL', formats['rbm_total_label'])
                    rbm_ws.write(total_row, 1, rbm_data['FTD Count'].sum(), formats['rbm_total'])
                    rbm_ws.write(total_row, 2, rbm_data['FTD Value'].sum(), formats['rbm_total'])
                    rbm_total_ftd_conversion = round((rbm_data['FTD Value'].sum() / rbm_data['Product_FTD_Amount'].sum()) * 100, 2) if rbm_data['Product_FTD_Amount'].sum() != 0 else 0
                    rbm_ws.write(total_row, 3, rbm_total_ftd_conversion / 100, formats['rbm_conversion_low'] if rbm_total_ftd_conversion < 2 else formats['rbm_total'])
                    rbm_ws.write(total_row, 4, rbm_data['MTD Count'].sum(), formats['rbm_total'])
                    rbm_ws.write(total_row, 5, rbm_data['MTD Value'].sum(), formats['rbm_total'])
                    rbm_total_mtd_conversion = round((rbm_data['MTD Value'].sum() / rbm_data['Product_MTD_Amount'].sum()) * 100, 2) if rbm_data['Product_MTD_Amount'].sum() != 0 else 0
                    rbm_ws.write(total_row, 6, rbm_total_mtd_conversion / 100, formats['rbm_conversion_low'] if rbm_total_mtd_conversion < 2 else formats['rbm_total'])
                    rbm_ws.write(total_row, 7, rbm_data['PREV MONTH SALE'].sum(), formats['rbm_total'])
                    total_prev = rbm_data['PREV MONTH SALE'].sum()
                    total_curr = rbm_data['MTD Value'].sum()
                    overall_growth = round(((total_curr - total_prev) / total_prev) * 100, 2) if total_prev != 0 else 0
                    rbm_ws.write(total_row, 8, f"{overall_growth}%", formats['rbm_total'])
                    overall_asp = round(total_curr / rbm_data['MTD Count'].sum(), 2) if rbm_data['MTD Count'].sum() != 0 else 0
                    rbm_ws.write(total_row, 9, overall_asp, formats['asp_total'])

                    insights_row = total_row + 2
                    if overall_growth > 15:
                        rbm_ws.merge_range(insights_row, 0, insights_row, len(headers) - 1,
                                         f"📈 Excellent Growth: {overall_growth}% increase from previous month",
                                         formats['rbm_summary'])
                    elif overall_growth < 0:
                        rbm_ws.merge_range(insights_row, 0, insights_row, len(headers) - 1,
                                         f"📉 Needs Attention: {abs(overall_growth)}% decrease from previous month",
                                         formats['rbm_summary'])
                    else:
                        rbm_ws.merge_range(insights_row, 0, insights_row, len(headers) - 1,
                                         f"📊 Stable Performance: Less change from previous month",
                                         formats['rbm_summary'])

                    insights_row += 1
                    top_3_stores = rbm_data.head(3)
                    if len(top_3_stores) > 0:
                        top_stores_text = " | ".join([f"{store['Store Name']}: ₹{int(store['MTD Value']):,}"
                                                    for _, store in top_3_stores.iterrows()])
                        rbm_ws.merge_range(insights_row, 0, insights_row, len(headers) - 1,
                                         f"🏆 Top 3 Performers: {top_stores_text}",
                                         formats['rbm_summary'])

            excel_output.seek(0)
            st.success("✅ Excel report generated successfully!")
            st.download_button(
                label="📥 Download Detailed Excel Report",
                data=excel_output.getvalue(),
                file_name=f"OSG_Sales_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Click to download the comprehensive sales report with all RBM sheets"
            )

    else:
        st.info("ℹ️ Please upload both Current Month OSG Sales Data and Current Month Product Sales Data to generate the report.")

# --------------------------- REPORT 2 TAB ---------------------------
with tab2:
    st.markdown('<h1 class="header">OSG Day View Report</h1>', unsafe_allow_html=True)

    with st.container():
        st.markdown("""
        <div class="info-box">
            <strong>Instructions:</strong> Upload the following file to generate the store summary report:
            <ul>
                <li><strong>Daily Sales Report</strong></li>
                <li><strong>myG Future Store List</strong> is loaded by default</li>
            </ul>
        </div>
        """, unsafe_allow_html=True)

    # Date and Time selection
    selected_date = st.date_input("Select Date", value=datetime.today())
    time_slot = st.selectbox("Select Time Slot", options=["12:30PM", "1PM", "4PM", "6PM"])
    formatted_date = selected_date.strftime("%d-%m-%Y")
    report_title = f"{formatted_date} EW Sale Till {time_slot}"

    # File uploader for sales report
    book2_file = st.file_uploader("Upload Daily Sales Report", type=["xlsx"], key="r2_book1")

    # Load Future Store List
    future_df = pd.read_excel("Future Store List.xlsx")
    st.success("✅ Loaded default Future Store List.")

    if book2_file:
        with st.spinner('Processing data...'):
            book2_df = pd.read_excel(book2_file)
            book2_df.rename(columns={'Branch': 'Store'}, inplace=True)

            agg = book2_df.groupby('Store', as_index=False).agg({
                'QUANTITY': 'sum',
                'AMOUNT': 'sum'
            })

            all_stores = pd.DataFrame(pd.concat([future_df['Store'], agg['Store']]).unique(), columns=['Store'])
            merged = all_stores.merge(agg, on='Store', how='left')
            merged['QUANTITY'] = merged['QUANTITY'].fillna(0).astype(int)
            merged['AMOUNT'] = merged['AMOUNT'].fillna(0).astype(int)

            merged = merged.sort_values(by='AMOUNT', ascending=False).reset_index(drop=True)

            total = pd.DataFrame([{
                'Store': 'TOTAL',
                'QUANTITY': merged['QUANTITY'].sum(),
                'AMOUNT': merged['AMOUNT'].sum()
            }])

            final_df = pd.concat([merged, total], ignore_index=True)
            final_df.rename(columns={'Store': 'Branch'}, inplace=True)

            # Excel report generator
            def generate_report2_excel(df, title_text):
                wb = Workbook()
                ws = wb.active
                ws.title = "Store Report"

                # Title
                ws.merge_cells('A1:C1')
                title_cell = ws['A1']
                title_cell.value = title_text
                title_cell.font = Font(bold=True, size=11, color="FFFFFF")
                title_cell.alignment = Alignment(horizontal='center')
                title_cell.fill = PatternFill("solid", fgColor="4F81BD")

                # Styles
                header_fill = PatternFill("solid", fgColor="4F81BD")
                data_fill = PatternFill("solid", fgColor="DCE6F1")
                red_fill = PatternFill("solid", fgColor="F4CCCC")
                total_fill = PatternFill("solid", fgColor="FFD966")
                border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                top=Side(style='thin'), bottom=Side(style='thin'))
                header_font = Font(bold=True, color="FFFFFF")
                bold_font = Font(bold=True)

                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=2):
                    for c_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=r_idx, column=c_idx, value=value)

                        if r_idx == 2:
                            cell.fill = header_fill
                            cell.font = header_font
                        elif df.loc[r_idx - 3, 'Branch'] == 'TOTAL':
                            cell.fill = total_fill
                            cell.font = bold_font
                        elif df.loc[r_idx - 3, 'AMOUNT'] <= 0:
                            cell.fill = red_fill
                        else:
                            cell.fill = data_fill

                        cell.border = border
                        cell.alignment = Alignment(horizontal='center')

                # Adjust column widths
                for col_idx, column_cells in enumerate(ws.columns, start=1):
                    max_length = 0
                    for cell in column_cells:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

                buffer = BytesIO()
                wb.save(buffer)
                buffer.seek(0)
                return buffer

            excel_buf2 = generate_report2_excel(final_df, report_title)

        with st.container():
            st.download_button(
                label="📥 Download Store Summary Report",
                data=excel_buf2,
                file_name=f"Store_Summary_{formatted_date}_{time_slot}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Download store summary report in Excel format"
            )
    else:
        st.info("ℹ️ Please upload the Daily Sales Report to generate the store summary.")

# --------------------------- REPORT 3 TAB ---------------------------
with tab3:
    st.markdown('<h1 class="header">OSG & Product Data Mapping</h1>', unsafe_allow_html=True)

    with st.container():
        st.markdown("""
        <div class="info-box">
            <strong>Instructions:</strong> Upload the following files to map OSG and product data:
            <ul>
                <li><strong>OSG File</strong> - Contains warranty and protection plan data</li>
                <li><strong>PRODUCT File</strong> - Contains product information including models, categories, and IMEIs</li>
            </ul>
        </div>
        """, unsafe_allow_html=True)

    # File upload section
    with st.container():
        st.markdown('<div class="file-upload-section">', unsafe_allow_html=True)
        osg_file = st.file_uploader(
            "Upload OSG File",
            type=["xlsx"],
            key="osg_mapping"
        )
        product_file = st.file_uploader(
            "Upload PRODUCT File",
            type=["xlsx"],
            key="product_mapping"
        )
        st.markdown('</div>', unsafe_allow_html=True)

    if osg_file and product_file:
        with st.spinner('Mapping data...'):
            osg_df = pd.read_excel(osg_file)
            product_df = pd.read_excel(product_file, converters={'IMEI': str})

            # SKU Mapping
            sku_category_mapping = {
                "Warranty : Water Cooler/Dispencer/Geyser/RoomCooler/Heater": [
                    "COOLER", "DISPENCER", "GEYSER", "ROOM COOLER", "HEATER", "WATER HEATER", "WATER DISPENSER"
                ],
                "Warranty : Fan/Mixr/IrnBox/Kettle/OTG/Grmr/Geysr/Steamr/Inductn": [
                    "FAN", "MIXER", "IRON BOX", "KETTLE", "OTG", "GROOMING KIT", "GEYSER", "STEAMER", "INDUCTION",
                    "CEILING FAN", "TOWER FAN", "PEDESTAL FAN", "INDUCTION COOKER", "ELECTRIC KETTLE", "WALL FAN", "MIXER GRINDER", "CELLING FAN"
                ],
                "AC : EWP : Warranty : AC": ["AC", "AIR CONDITIONER", "AC INDOOR"],
                "HAEW : Warranty : Air Purifier/WaterPurifier": ["AIR PURIFIER", "WATER PURIFIER"],
                "HAEW : Warranty : Dryer/MW/DishW": ["DRYER", "MICROWAVE OVEN", "DISH WASHER", "MICROWAVE OVEN-CONV"],
                "HAEW : Warranty : Ref/WM": [
                    "REFRIGERATOR", "WASHING MACHINE", "WASHING MACHINE-TL", "REFRIGERATOR-DC",
                    "WASHING MACHINE-FL", "WASHING MACHINE-SA", "REF", "REFRIGERATOR-CBU", "REFRIGERATOR-FF", "WM"
                ],
                "HAEW : Warranty : TV": ["TV", "TV 28 %", "TV 18 %"],
                "TV : TTC : Warranty and Protection : TV": ["TV", "TV 28 %", "TV 18 %"],
                "TV : Spill and Drop Protection": ["TV", "TV 28 %", "TV 18 %"],
                "HAEW : Warranty :Chop/Blend/Toast/Air Fryer/Food Processr/JMG/Induction": [
                    "CHOPPER", "BLENDER", "TOASTER", "AIR FRYER", "FOOD PROCESSOR", "JUICER", "INDUCTION COOKER"
                ],
                "HAEW : Warranty : HOB and Chimney": ["HOB", "CHIMNEY"],
                "HAEW : Warranty : HT/SoundBar/AudioSystems/PortableSpkr": [
                    "HOME THEATRE", "AUDIO SYSTEM", "SPEAKER", "SOUND BAR", "PARTY SPEAKER"
                ],
                "HAEW : Warranty : Vacuum Cleaner/Fans/Groom&HairCare/Massager/Iron": [
                    "VACUUM CLEANER", "FAN", "MASSAGER", "IRON BOX", "CEILING FAN", "TOWER FAN", "PEDESTAL FAN", "WALL FAN", "ROBO VACCUM CLEANER"
                ],
                "AC AMC": ["AC", "AC INDOOR"]
            }

            product_df['Category'] = product_df['Category'].str.upper().fillna('')
            product_df['Model'] = product_df['Model'].fillna('')
            product_df['Customer Mobile'] = product_df['Customer Mobile'].astype(str)
            product_df['Invoice Number'] = product_df['Invoice Number'].astype(str)
            product_df['Item Rate'] = pd.to_numeric(product_df['Item Rate'], errors='coerce')
            product_df['IMEI'] = product_df['IMEI'].astype(str).fillna('')
            product_df['Brand'] = product_df['Brand'].fillna('')
            osg_df['Customer Mobile'] = osg_df['Customer Mobile'].astype(str)

            def extract_price_slab(text):
                match = re.search(r"Slab\s*:\s*(\d+)K-(\d+)K", str(text))
                if match:
                    return int(match.group(1)) * 1000, int(match.group(2)) * 1000
                return None, None

            def get_model(row):
                mobile = row['Customer Mobile']
                retailer_sku = str(row['Retailer SKU'])
                invoice = str(row.get('Invoice Number', ''))
                user_products = product_df[product_df['Customer Mobile'] == mobile]

                if user_products.empty:
                    return ''
                unique_models = user_products['Model'].dropna().unique()
                if len(unique_models) == 1:
                    return unique_models[0]

                mapped_keywords = []
                for sku_key, keywords in sku_category_mapping.items():
                    if sku_key in retailer_sku:
                        mapped_keywords = [kw.lower() for kw in keywords]
                        break   

                filtered = user_products[user_products['Category'].str.lower().isin(mapped_keywords)]
                if filtered['Model'].nunique() == 1:
                    return filtered['Model'].iloc[0]

                slab_min, slab_max = extract_price_slab(retailer_sku)
                if slab_min and slab_max:
                    slab_filtered = filtered[(filtered['Item Rate'] >= slab_min) & (filtered['Item Rate'] <= slab_max)]
                    if slab_filtered['Model'].nunique() == 1:
                        return slab_filtered['Model'].iloc[0]
                    invoice_filtered = slab_filtered[slab_filtered['Invoice Number'].astype(str) == invoice]
                    if invoice_filtered['Model'].nunique() == 1:
                        return invoice_filtered['Model'].iloc[0]

                return ''

            osg_df['Model'] = osg_df.apply(get_model, axis=1)
            category_brand_df = product_df[['Customer Mobile', 'Model', 'Category', 'Brand']].drop_duplicates()
            osg_df = osg_df.merge(category_brand_df, on=['Customer Mobile', 'Model'], how='left')

            invoice_pool = defaultdict(list)
            itemrate_pool = defaultdict(list)
            imei_pool = defaultdict(list)

            for _, row in product_df.iterrows():
                key = (row['Customer Mobile'], row['Model'])
                invoice_pool[key].append(row['Invoice Number'])
                itemrate_pool[key].append(row['Item Rate'])
                imei_pool[key].append(row['IMEI'])

            invoice_usage_counter = defaultdict(int)
            itemrate_usage_counter = defaultdict(int)
            imei_usage_counter = defaultdict(int)

            def assign_from_pool(row, pool, counter_dict):
                key = (row['Customer Mobile'], row['Model'])
                values = pool.get(key, [])
                index = counter_dict[key]
                if index < len(values):
                    counter_dict[key] += 1
                    return values[index]
                return ''

            osg_df['Product Invoice Number'] = osg_df.apply(lambda row: assign_from_pool(row, invoice_pool, invoice_usage_counter), axis=1)
            osg_df['Item Rate'] = osg_df.apply(lambda row: assign_from_pool(row, itemrate_pool, itemrate_usage_counter), axis=1)
            osg_df['IMEI'] = osg_df.apply(lambda row: assign_from_pool(row, imei_pool, imei_usage_counter), axis=1)
            osg_df['Store Code'] = osg_df['Product Invoice Number'].astype(str).apply(
                lambda x: re.search(r'\b([A-Z]{2,})\b', x).group(1) if re.search(r'\b([A-Z]{2,})\b', x) else ''
            )

            def extract_warranty_duration(sku):
                sku = str(sku)
                match = re.search(r'Dur\s*:\s*(\d+)\+(\d+)', sku)
                if match:
                    return int(match.group(1)), int(match.group(2))
                match = re.search(r'(\d+)\+(\d+)\s*SDP-(\d+)', sku)
                if match:
                    return int(match.group(1)), f"{match.group(3)}P+{match.group(2)}W"
                match = re.search(r'Dur\s*:\s*(\d+)', sku)
                if match:
                    return 1, int(match.group(1))
                match = re.search(r'(\d+)\+(\d+)', sku)
                if match:
                    return int(match.group(1)), int(match.group(2))
                return '', ''

            osg_df[['Manufacturer Warranty', 'Duration (Year)']] = osg_df['Retailer SKU'].apply(
                lambda sku: pd.Series(extract_warranty_duration(sku))
            )

            def highlight_row(row):
                missing_fields = pd.isna(row.get('Model')) or str(row.get('Model')).strip() == ''
                missing_fields |= pd.isna(row.get('IMEI')) or str(row.get('IMEI')).strip() == ''
                try:
                    if float(row.get('Plan Price', 0)) < 0:
                        missing_fields |= True
                except:
                    missing_fields |= True
                return ['background-color: lightblue'] * len(row) if missing_fields else [''] * len(row)

            final_columns = [
                'Customer Mobile', 'Date', 'Invoice Number','Product Invoice Number', 'Customer Name', 'Store Code', 'Branch', 'Region',
                'IMEI', 'Category', 'Brand', 'Quantity', 'Item Code', 'Model', 'Plan Type', 'EWS QTY', 'Item Rate',
                'Plan Price', 'Sold Price', 'Email', 'Product Count', 'Manufacturer Warranty', 'Retailer SKU', 'OnsiteGo SKU',
                'Duration (Year)', 'Total Coverage', 'Comment', 'Return Flag', 'Return against invoice No.',
                'Primary Invoice No.'
            ]

            for col in final_columns:
                if col not in osg_df.columns:
                    osg_df[col] = ''
            osg_df['Quantity'] = 1
            osg_df['EWS QTY'] = 1
            osg_df = osg_df[final_columns]

            st.markdown("""
            <div class="success-box">
                <strong>✅ Data Mapping Completed Successfully</strong>
                <p>The OSG and product data has been successfully mapped. You can now download the report.</p>
            </div>
            """, unsafe_allow_html=True)

            @st.cache_data
            def convert_df(df):
               output = io.BytesIO()
               styled_df = df.style.apply(highlight_row, axis=1)
               with pd.ExcelWriter(output, engine='openpyxl') as writer:
                styled_df.to_excel(writer, index=False)
               output.seek(0)
               return output

            excel_data = convert_df(osg_df)

        # Download section
        with st.container():
            st.markdown('<div class="download-section">', unsafe_allow_html=True)
            st.download_button(
                label="📥 Download Mapped Data Report",
                data=excel_data,
                file_name="OSG_Product_Mapping_Report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Download the mapped OSG and product data in Excel format"
            )
            st.markdown('</div>', unsafe_allow_html=True)
    else:
        st.info("ℹ️ Please upload both required files to perform data mapping.")
